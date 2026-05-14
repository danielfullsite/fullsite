"""Export handler — download XLSX files and extract schema info."""

import asyncio
import json
import os

from openpyxl import load_workbook

from .crawler import MenuItem, slugify


async def download_exports(page, items: list[MenuItem], output_dir: str = "output") -> dict:
    """
    For each item with has_export=True, click the export button,
    download the XLSX, and extract schema info.

    Loads previous schemas to skip already-downloaded exports (resume support).
    Returns merged xlsx_schemas dict.
    """
    xlsx_dir = os.path.join(output_dir, "xlsx_samples")
    os.makedirs(xlsx_dir, exist_ok=True)

    # Load previously downloaded schemas for resume
    schemas = {}
    schemas_path = os.path.join(output_dir, "xlsx_schemas.json")
    if os.path.exists(schemas_path):
        with open(schemas_path, "r", encoding="utf-8") as f:
            schemas = json.load(f)

    export_items = [item for item in items if item.has_export and item.href]

    # Filter to only pending exports
    pending = []
    for item in export_items:
        slug = slugify(item.path.replace("/", "-"))
        file_path = os.path.join(xlsx_dir, f"{slug}.xlsx")
        if item.path in schemas or os.path.exists(file_path):
            continue  # Already downloaded
        pending.append(item)

    print(f"\n[export] === Paso 5: Exports ===")
    print(f"[export] Total with export: {len(export_items)}")
    print(f"[export] Already downloaded: {len(export_items) - len(pending)}")
    print(f"[export] Pending: {len(pending)}")

    if not pending:
        print("[export] Nothing to download")
        return schemas

    for i, item in enumerate(pending):
        print(f"[export] ({i + 1}/{len(pending)}) {item.path}")

        try:
            # Navigate to the page
            url = item.href
            if not url.startswith("http"):
                url = "https://www.wansoft.net/Wansoft.Web" + url

            await page.goto(url, wait_until="domcontentloaded", timeout=15000)
            await asyncio.sleep(2)

            # Select all sucursales in any multi-select (fixes "Favor de seleccionar" modal)
            await _select_all_subsidiaries(page)

            # Try to set date filters
            await _try_set_date_filters(page)

            # Clear any overlays or modals that may be blocking
            await _dismiss_modals(page)

            # Find the export button
            export_btn = await _find_export_button(page)
            if not export_btn:
                print(f"[export]   SKIP: no export button found")
                item.has_export = False
                continue

            # Download the file
            slug = slugify(item.path.replace("/", "-"))
            file_path = os.path.join(xlsx_dir, f"{slug}.xlsx")

            try:
                async with page.expect_download(timeout=20000) as download_info:
                    await export_btn.click(force=True)
                download = await download_info.value
                await download.save_as(file_path)
                print(f"[export]   Downloaded: {file_path}")
            except Exception as e:
                # Check if modal appeared (sucursal not selected)
                modal_dismissed = await _dismiss_modals(page)
                if modal_dismissed:
                    # Retry after dismissing modal and selecting sucursal
                    await _select_all_subsidiaries(page)
                    await asyncio.sleep(1)
                    try:
                        async with page.expect_download(timeout=20000) as download_info:
                            await export_btn.click(force=True)
                        download = await download_info.value
                        await download.save_as(file_path)
                        print(f"[export]   Downloaded (retry): {file_path}")
                    except Exception as e2:
                        print(f"[export]   DOWNLOAD FAILED: {e2}")
                        item.notes += f" Export failed: {e2}"
                        continue
                else:
                    print(f"[export]   DOWNLOAD FAILED: {e}")
                    item.notes += f" Export failed: {e}"
                    continue

            # Parse XLSX schema
            schema = _parse_xlsx_schema(file_path)
            if schema:
                item.xlsx_sample_path = f"xlsx_samples/{slug}.xlsx"
                schemas[item.path] = {
                    "file_path": item.xlsx_sample_path,
                    "sheets": schema,
                }
                cols_count = sum(len(s.get("columns", [])) for s in schema)
                print(f"[export]   Schema: {len(schema)} sheets, {cols_count} total columns")

            await asyncio.sleep(2)

        except Exception as e:
            print(f"[export]   ERROR: {e}")
            item.notes += f" Export error: {e}"
            continue

    # Save merged schemas
    with open(schemas_path, "w", encoding="utf-8") as f:
        json.dump(schemas, f, indent=2, ensure_ascii=False)
    print(f"\n[export] Saved {len(schemas)} total schemas to {schemas_path}")

    return schemas


async def _select_all_subsidiaries(page):
    """Select all options in any multi-select (sucursales) on the page."""
    await page.evaluate('''() => {
        const selects = document.querySelectorAll('select[multiple]');
        selects.forEach(sel => {
            let changed = false;
            for (const opt of sel.options) {
                if (!opt.selected) {
                    opt.selected = true;
                    changed = true;
                }
            }
            if (changed) {
                sel.dispatchEvent(new Event('change', {bubbles: true}));
            }
        });
    }''')
    await asyncio.sleep(0.5)


async def _dismiss_modals(page) -> bool:
    """Dismiss any visible modals/overlays. Returns True if a modal was found."""
    dismissed = await page.evaluate('''() => {
        let found = false;

        // Remove jQuery UI overlays
        document.querySelectorAll('.ui-widget-overlay').forEach(e => {
            e.remove();
            found = true;
        });

        // Remove Bootstrap modal backdrops
        document.querySelectorAll('.modal-backdrop').forEach(e => {
            e.remove();
            found = true;
        });

        // Close visible modals by clicking close/accept buttons
        const modalSelectors = ['.modal.in', '.modal.show', '.modal[style*="block"]', '.bootbox', '.ui-dialog'];
        for (const sel of modalSelectors) {
            const modal = document.querySelector(sel);
            if (modal) {
                // Try clicking Aceptar, OK, or Close
                const btns = modal.querySelectorAll('button, a.btn');
                for (const btn of btns) {
                    const text = btn.textContent.trim().toLowerCase();
                    if (text === 'aceptar' || text === 'ok' || text === 'cerrar' || text === 'close') {
                        btn.click();
                        found = true;
                        break;
                    }
                }
                // Also try the X close button
                const closeBtn = modal.querySelector('.close, .ui-dialog-titlebar-close');
                if (closeBtn) closeBtn.click();

                // Hide the modal
                modal.style.display = 'none';
                modal.classList.remove('in', 'show');
                found = true;
            }
        }

        // Re-enable body scroll
        document.body.classList.remove('modal-open');
        document.body.style.overflow = '';

        return found;
    }''')
    if dismissed:
        await asyncio.sleep(0.5)
    return dismissed


async def _find_export_button(page):
    """Find the export button on the current page."""
    export_selectors = [
        '#btnExport', 'input[value*="Exportar"]', 'input[value*="Excel"]',
        'a:has-text("Excel")', 'a:has-text("Exportar")',
        'button:has-text("Excel")', 'button:has-text("Exportar")',
        '.btn-export', 'a[href*="Export"]',
        'a:has-text("Descargar")', 'button:has-text("Descargar")',
    ]
    for esel in export_selectors:
        try:
            eel = page.locator(esel).first
            if await eel.is_visible(timeout=500):
                return eel
        except Exception:
            continue
    return None


async def _try_set_date_filters(page):
    """Try to set date filters to today or last 7 days."""
    try:
        date_inputs = page.locator('input[type="date"], input[name*="Date" i], input[name*="fecha" i]')
        count = await date_inputs.count()
        if count > 0:
            from datetime import date, timedelta
            today = date.today().isoformat()
            week_ago = (date.today() - timedelta(days=7)).isoformat()

            if count >= 2:
                try:
                    await date_inputs.nth(0).fill(week_ago)
                    await date_inputs.nth(1).fill(today)
                except Exception:
                    pass
            elif count == 1:
                try:
                    await date_inputs.nth(0).fill(today)
                except Exception:
                    pass
    except Exception:
        pass


def _parse_xlsx_schema(file_path: str) -> list[dict] | None:
    """Extract schema from an XLSX file."""
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(max_row=4, values_only=True))
            if not rows:
                continue

            columns = [str(c) if c is not None else "" for c in rows[0]]
            columns = [c for c in columns if c]

            sample_row = []
            if len(rows) > 1:
                sample_row = [_serialize_cell(c) for c in rows[1][:len(columns)]]

            sheets.append({
                "name": sheet_name,
                "columns": columns,
                "sample_row": sample_row,
            })

        wb.close()
        return sheets if sheets else None
    except Exception as e:
        print(f"[export]   XLSX parse error: {e}")
        return None


def _serialize_cell(value):
    """Convert cell value to JSON-serializable form."""
    if value is None:
        return None
    if isinstance(value, (int, float, str, bool)):
        return value
    return str(value)
