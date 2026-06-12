#!/usr/bin/env python3
"""Reconcile fresh recetas.xlsx (2026-06-12) vs wansoft_recipes (06-10) vs 522 active platillos."""
import json
import sys
import unicodedata
from openpyxl import load_workbook


def norm(s):
    s = unicodedata.normalize('NFD', str(s or ''))
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return ' '.join(s.upper().split())


def parse_xlsx(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # find header
    hi = None
    for i, r in enumerate(rows):
        if r and any('Platillo' in str(c or '') for c in r) and any('Producto' in str(c or '') for c in r):
            hi = i
            break
    hdr = [str(c or '').strip() for c in rows[hi]]
    ix = {h: j for j, h in enumerate(hdr)}
    recipes = {}  # norm(platillo) -> {'codigo':..., 'name':..., 'ings': {norm(producto): (cantidad, unidad)}}
    for r in rows[hi + 1:]:
        if not r or not r[ix['Platillo']] or not r[ix['Producto']]:
            continue
        plat = norm(r[ix['Platillo']])
        prod = norm(r[ix['Producto']])
        qty = r[ix['Cantidad']]
        unit = str(r[ix['Unidad de medida']] or '').strip()
        cod = str(r[ix['Código del platillo']] or '').strip()
        e = recipes.setdefault(plat, {'codigo': cod, 'name': str(r[ix['Platillo']]).strip(), 'ings': {}})
        e['ings'][prod] = (float(qty or 0), unit)
    return recipes


def parse_db(path):
    data = json.load(open(path))
    recipes = {}
    for row in data:
        plat = norm(row['saucer_name'])
        ings = {}
        for ing in (row.get('ingredients') or []):
            ings[norm(ing.get('ProductName'))] = (
                float(ing.get('Quantity') or 0),
                str(ing.get('UnitOfMeasureDescription') or '').strip(),
            )
        recipes[plat] = {'name': row['saucer_name'], 'ings': ings, 'saucer_id': row.get('saucer_id')}
    return recipes


def main():
    fresh = parse_xlsx(sys.argv[1])
    db = parse_db(sys.argv[2])
    items = json.load(open(sys.argv[3]))
    active = {norm(i['name']): i for i in items if i.get('active')}

    print(f"Fresh XLSX: {len(fresh)} platillos con receta")
    print(f"DB wansoft_recipes (06-10): {len(db)} platillos")
    print(f"POS activos: {len(active)} items")
    print()

    fk, dk = set(fresh), set(db)
    added = sorted(fk - dk)
    removed = sorted(dk - fk)
    changed = []
    for k in fk & dk:
        a, b = fresh[k]['ings'], db[k]['ings']
        if set(a) != set(b):
            changed.append((k, 'ingredientes', sorted(set(a) ^ set(b))))
        else:
            diffs = [(p, b[p][0], a[p][0]) for p in a if abs(a[p][0] - b[p][0]) > 1e-9]
            if diffs:
                changed.append((k, 'cantidades', diffs))

    print(f"=== DRIFT fresh vs DB ===")
    print(f"Recetas nuevas (en XLSX, no en DB): {len(added)}")
    for k in added[:30]:
        print(f"  + {fresh[k]['name']} ({len(fresh[k]['ings'])} ings)")
    if len(added) > 30:
        print(f"  ... y {len(added)-30} más")
    print(f"Recetas removidas (en DB, no en XLSX): {len(removed)}")
    for k in removed[:30]:
        print(f"  - {db[k]['name']}")
    if len(removed) > 30:
        print(f"  ... y {len(removed)-30} más")
    print(f"Recetas modificadas: {len(changed)}")
    for k, kind, det in changed[:30]:
        print(f"  ~ {fresh[k]['name']} ({kind}): {det[:6]}")
    print()

    # Cross vs active POS items
    con_receta = set(active) & fk
    sin_receta = sorted(set(active) - fk)
    huerfanas = sorted(fk - set(active))
    print(f"=== COBERTURA FOOD COST (522 activos) ===")
    print(f"Activos CON receta: {len(con_receta)}")
    print(f"Activos SIN receta: {len(sin_receta)}")
    # group sin receta by category
    bycat = {}
    for k in sin_receta:
        bycat.setdefault(active[k]['category_id'], []).append(active[k]['name'])
    for cat in sorted(bycat, key=lambda c: -len(bycat[c])):
        names = bycat[cat]
        print(f"  [{cat}] {len(names)}: {', '.join(names[:8])}{' ...' if len(names) > 8 else ''}")
    print(f"Recetas huérfanas (XLSX sin item activo): {len(huerfanas)}")
    for k in huerfanas[:20]:
        print(f"  ? {fresh[k]['name']}")
    if len(huerfanas) > 20:
        print(f"  ... y {len(huerfanas)-20} más")


if __name__ == '__main__':
    main()
