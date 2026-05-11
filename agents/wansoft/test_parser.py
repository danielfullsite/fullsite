#!/usr/bin/env python3
"""Generate a synthetic test XLSX and validate parser output."""

import sys
from pathlib import Path

import openpyxl

TEST_DIR = Path(__file__).parent / "test_data"
TEST_FILE = TEST_DIR / "ReporteVentasPorMesero2026-05-10.xlsx"

# Data approximating the 09-May demo numbers
MESEROS_DATA = [
    # (Mesero, Subtotal, IEPS, IVA, Total, %, Personas, Promedio)
    ("Brayan Berlanga Solis",       14200, 0, 2272, 16472, 15.7, 26, 623),
    ("Oscar Rios Alvarado",         12800, 0, 2048, 14848, 14.1, 29, 506),
    ("Julio Cesar Hernández H.",    11000, 0, 1760, 12760, 12.1, 26, 482),
    ("Omar Aguilera",               13500, 0, 2160, 15660, 14.9, 35, 443),
    ("Hector Enrique Rodriguez L.", 10200, 0, 1632, 11832, 11.3, 28, 419),
    ("Daniela Edith Rico Segura",    8500, 0, 1360,  9860,  9.4, 24, 408),
    ("Mauricio Rodriguez Rodriguez", 7800, 0, 1248,  9048,  8.6, 23, 390),
    ("Alexis Alejandro Ocampo V.",   5100, 0,  816,  5916,  5.6, 16, 365),
    ("MESERO EVENTO",                4200, 0,  672,  4872,  4.6, 12, 406),
    ("APLICACIONES",                 3825, 0,  612,  4437,  4.2,  0,   0),
]

# Granular platillo data for Sheet 1 (Ventas por mesero por grupo)
# (Fecha, Mesero, TipoGrupo, Grupo, Platillo, Cantidad, Subtotal, IVA, IEPS, Total, %)
DETALLE_DATA = [
    # Brayan — H&H, chilaquiles, bakery, desserts
    ("2026-05-10", "Brayan Berlanga Solis", "ALIMENTOS", "SIGNATURE", "HALF  HALF COMBO", 5, 1163, 186, "", 1350, 8.2),
    ("2026-05-10", "Brayan Berlanga Solis", "ALIMENTOS", "SIGNATURE", "HALF  HALF COMBO", 1, 232, 37, "", 270, 1.6),
    ("2026-05-10", "Brayan Berlanga Solis", "ALIMENTOS", "CHILAQUILES & ENCHILADAS", "CHILAQUILES", 3, 646, 103, "", 750, 4.6),
    ("2026-05-10", "Brayan Berlanga Solis", "ALIMENTOS", "CHILAQUILES & ENCHILADAS", "CHILAQUILES LIGHT", 1, 215, 34, "", 250, 1.5),
    ("2026-05-10", "Brayan Berlanga Solis", "POSTRES", "BAKERY", "CROISSANT NUTELLA", 3, 256, 41, "", 297, 1.8),
    ("2026-05-10", "Brayan Berlanga Solis", "POSTRES", "BAKERY", "CINAMMON ROLL", 1, 85, 13, "", 99, 0.6),
    ("2026-05-10", "Brayan Berlanga Solis", "POSTRES", "DESSERTS", "NEW YORK CHEESECAKE", 2, 224, 35, "", 260, 1.6),
    # Oscar — H&H, bakery, ice cream
    ("2026-05-10", "Oscar Rios Alvarado", "ALIMENTOS", "SIGNATURE", "HALF  HALF COMBO", 2, 465, 74, "", 540, 3.6),
    ("2026-05-10", "Oscar Rios Alvarado", "POSTRES", "BAKERY", "CHOCOLATIN", 2, 118, 18, "", 138, 0.9),
    ("2026-05-10", "Oscar Rios Alvarado", "POSTRES", "ICE CREAM", "NIEVE FRUTOS ROJOS", 1, 64, 10, "", 75, 0.5),
    # Omar — chilaquiles, bakery
    ("2026-05-10", "Omar Aguilera", "ALIMENTOS", "CHILAQUILES & ENCHILADAS", "CHILAQUILES", 7, 1508, 241, "", 1750, 11.2),
    ("2026-05-10", "Omar Aguilera", "POSTRES", "BAKERY", "CROISSANT NUTELLA", 5, 427, 68, "", 495, 3.0),
    # Hector — bakery, desserts
    ("2026-05-10", "Hector Enrique Rodriguez L.", "POSTRES", "BAKERY", "SUNSHINE MUFFIN", 2, 167, 26, "", 194, 1.2),
    ("2026-05-10", "Hector Enrique Rodriguez L.", "POSTRES", "DESSERTS", "DARK CHOCOLATE BROWNIE", 3, 336, 53, "", 390, 2.5),
    # MESERO EVENTO — should be excluded from per-mesero breakdowns
    ("2026-05-10", "MESERO EVENTO", "POSTRES", "BAKERY", "CROISSANT NUTELLA", 10, 854, 136, "", 990, 6.0),
    ("2026-05-10", "MESERO EVENTO", "POSTRES", "DESSERTS", "CARROT CAKE", 4, 465, 74, "", 540, 3.3),
    # APLICACIONES — should be excluded
    ("2026-05-10", "APLICACIONES", "ALIMENTOS", "SIGNATURE", "HALF  HALF COMBO", 3, 697, 111, "", 810, 5.2),
    ("2026-05-10", "APLICACIONES", "POSTRES", "BAKERY", "CHOCOLATIN", 2, 118, 18, "", 138, 0.9),
]


def create_test_xlsx():
    """Create a synthetic XLSX matching real Wansoft export structure.

    Real structure: Sheet 1 = detalle por platillo, Sheet 2 = resumen por mesero.
    Both have 2 empty leading columns (A, B).
    """
    TEST_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # Sheet 1: Ventas por mesero por grupo (detalle)
    ws1 = wb.active
    ws1.title = "Ventas por mesero por grupo"
    # Metadata rows with 2 leading empty cols
    ws1.append(["", ""])
    ws1.append(["", "", "Sucursal: Café Amalay - Plaza Duendes"])
    ws1.append(["", "", "Reporte del: 2026-05-10 al 2026-05-10"])
    ws1.append(["", "", "Reporte generado: 2026-05-10"])
    ws1.append(["", "", "Total descuentos sobre cuentas:", "", "",
                "Reporte Ventas Por Mesero Por Grupo"])
    ws1.append(["", ""])  # blank
    # Header row (row 7) — matches real structure
    ws1.append(["", "", "Fecha", "Mesero", "Tipo grupo", "Grupo", "Platillo",
                "Cantidad", "Subtotal", "IVA", "IEPS", "Total", "%"])
    # Data rows
    for row in DETALLE_DATA:
        ws1.append(["", ""] + list(row))

    # Sheet 2: Resumen de ventas por mesero
    ws2 = wb.create_sheet("Resumen de ventas por mesero")
    ws2.append(["", ""])
    ws2.append(["", "", "Sucursal: Café Amalay - Plaza Duendes"])
    ws2.append(["", "", "Reporte del: 2026-05-10 al 2026-05-10"])
    ws2.append(["", "", "Reporte generado: 2026-05-10"])
    ws2.append(["", "", "Total descuentos sobre cuentas:", "", "",
                "Reporte ventas por mesero"])
    ws2.append(["", ""])  # blank
    # Header row (row 7)
    ws2.append(["", "", "Mesero", "Subtotal", "IEPS", "IVA", "Total", "%",
                "Personas atendidas", "Promedio por persona"])
    for row in MESEROS_DATA:
        ws2.append(["", ""] + list(row))
    ws2.append(["", "", "Total", 91125, 0, 14580, 105705, 100, 219, 0])

    # Sheet 3 & 4 (stubs)
    wb.create_sheet("Ventas por mesero por mesa")
    wb.create_sheet("Ventas por tipo de grupo")

    wb.save(TEST_FILE)
    return TEST_FILE


def test_format_message():
    xlsx = create_test_xlsx()

    sys.path.insert(0, str(Path(__file__).parent))
    from parser import format_message

    msg = format_message(str(xlsx))
    print("=== MESERO OUTPUT ===")
    print(msg)
    print("=== END ===\n")

    errors = []

    if "Cierre 10 May" not in msg:
        errors.append(f"Date: expected 'Cierre 10 May', got: {msg.splitlines()[0]}")

    if "APLICACIONES" in msg:
        errors.append("APLICACIONES should be filtered out")

    if "MESERO EVENTO" in msg:
        errors.append("MESERO EVENTO should be filtered out")

    lines = msg.splitlines()
    medal_lines = [l for l in lines if l.startswith(("🥇", "🥈", "🥉"))]
    if len(medal_lines) != 3:
        errors.append(f"Expected 3 medal lines, got {len(medal_lines)}")
    else:
        if "Brayan Berlanga" not in medal_lines[0]:
            errors.append(f"Top 1 should be Brayan, got: {medal_lines[0]}")
        if "$623" not in medal_lines[0]:
            errors.append(f"Top 1 avg should be $623, got: {medal_lines[0]}")
        if "Oscar Rios" not in medal_lines[1]:
            errors.append(f"Top 2 should be Oscar Rios, got: {medal_lines[1]}")
        if "Julio Cesar" not in medal_lines[2]:
            errors.append(f"Top 3 should be Julio Cesar, got: {medal_lines[2]}")

    if "Total día:" not in msg:
        errors.append("Missing 'Total día:' line")
    if "207 personas" not in msg:
        errors.append("Expected 207 personas in total line")
    if "$466" not in msg:
        errors.append("Expected general avg ~$466")

    return errors


def test_format_platillos():
    xlsx_path = str(TEST_FILE)

    sys.path.insert(0, str(Path(__file__).parent))
    from parser import format_platillos_message

    msg = format_platillos_message(xlsx_path)
    print("=== PLATILLOS OUTPUT ===")
    print(msg)
    print("=== END ===\n")

    errors = []

    # Header
    if "Detalle platillos 10 May" not in msg:
        errors.append(f"Header: expected 'Detalle platillos 10 May', got: {msg.splitlines()[0]}")

    # H&H total: Brayan 5+1=6, Oscar 2, APLICACIONES 3 (excluded from per-mesero
    # but H&H count is global across ALL rows including APLICACIONES)
    # Actually: H&H is sum of ALL rows with HALF HALF COMBO
    # Brayan: 5+1=6, Oscar: 2, APLICACIONES: 3 = 11 total
    # Wait — the spec says "sum(Cantidad) where Platillo contains HALF HALF COMBO"
    # It doesn't say to exclude APLICACIONES for the global count.
    # Let me re-read: "Filter out APLICACIONES and MESERO EVENTO from per-mesero breakdowns"
    # So global H&H and chilaquiles counts include ALL rows.
    hh_expected = 6 + 2 + 3  # Brayan + Oscar + APLICACIONES
    if f"{hh_expected} vendidos" not in msg.split("H&H")[1].split("\n")[0]:
        # Try checking the line
        hh_line = [l for l in msg.splitlines() if "H&H" in l]
        if hh_line and f"{hh_expected}" not in hh_line[0]:
            errors.append(f"H&H expected {hh_expected}, got: {hh_line[0]}")

    # Chilaquiles: Brayan 3+1=4, Omar 7 = 11
    chil_expected = 4 + 7
    chil_line = [l for l in msg.splitlines() if "Chilaquiles" in l]
    if chil_line and f"{chil_expected}" not in chil_line[0]:
        errors.append(f"Chilaquiles expected {chil_expected}, got: {chil_line[0]}")

    # Bakery per mesero (excluding MESERO EVENTO and APLICACIONES):
    # Omar: 5, Brayan: 3+1=4, Oscar: 2, Hector: 2
    # Total: 13
    if "MESERO EVENTO" in msg:
        errors.append("MESERO EVENTO should be excluded from per-mesero bakery")
    if "APLICACIONES" in msg:
        errors.append("APLICACIONES should be excluded from per-mesero bakery")

    bakery_total_line = [l for l in msg.splitlines() if "piezas" in l]
    # First "piezas" line should be bakery total
    if bakery_total_line:
        if "13 piezas" not in bakery_total_line[0]:
            errors.append(f"Bakery total expected 13, got: {bakery_total_line[0]}")

    # Postres per mesero (DESSERTS + ICE CREAM, excluding excluded meseros):
    # Brayan: 2 (cheesecake), Hector: 3 (brownie), Oscar: 1 (ice cream) = 6
    if len(bakery_total_line) >= 2:
        if "6 piezas" not in bakery_total_line[1]:
            errors.append(f"Postres total expected 6, got: {bakery_total_line[1]}")

    # Avance mode
    avance_msg = format_platillos_message(xlsx_path, report_type="avance")
    if "(3pm)" not in avance_msg:
        errors.append("Avance header should contain '(3pm)'")

    return errors


if __name__ == "__main__":
    all_errors = []
    all_errors.extend(test_format_message())
    all_errors.extend(test_format_platillos())

    if all_errors:
        print("FAILURES:")
        for e in all_errors:
            print(f"  ✗ {e}")
        sys.exit(1)
    else:
        print("All checks passed.")
