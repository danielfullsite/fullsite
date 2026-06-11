#!/usr/bin/env python3
"""Parse Wansoft "Reporte Ventas Por Mesero" XLSX and emit WhatsApp message."""

import re
import sys
import statistics
from datetime import date, datetime
from pathlib import Path

import openpyxl

EXCLUDE_MESEROS = {
    "APLICACIONES",
    "MESERO EVENTO",
    "Oscar Ricardo",   # Supervisor caja Take Away (no atiende mesa)
    "Hector Enrique",  # Cajero, confirmado por Mónica 2026-05-10
    "Rodrigo Chávez",   # Cajero, confirmado por Mónica 2026-05-11
    "Fany Elizabeth",   # Cajera, confirmada por Mónica 2026-05-11
}


def _is_excluded(mesero: str) -> bool:
    """Check if mesero matches any exclusion (substring, case-insensitive)."""
    m = mesero.upper()
    return any(ex.upper() in m for ex in EXCLUDE_MESEROS)

RESUMEN_SHEET = "Resumen de ventas por mesero"

MESES_ES = {
    1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Ago", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic",
}

MEDALS = ["🥇", "🥈", "🥉"]


def _short_name(full: str) -> str:
    """First two words of a name."""
    parts = full.strip().split()
    return " ".join(parts[:2])


def _fmt_money(val: float) -> str:
    """$1,234 — rounded peso, no decimals."""
    return f"${round(val):,}"


def _find_resumen_sheet(wb: openpyxl.Workbook):
    """Find the 'Resumen de ventas por mesero' sheet by name or fallback."""
    # Try exact match first
    if RESUMEN_SHEET in wb.sheetnames:
        return wb[RESUMEN_SHEET]
    # Case-insensitive search
    for name in wb.sheetnames:
        if "resumen" in name.lower() and "mesero" in name.lower():
            return wb[name]
    # Fallback: scan all sheets for the header pattern
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
            vals = [str(c).strip().lower() if c else "" for c in row]
            if "mesero" in vals and "promedio por persona" in " ".join(vals):
                return ws
    raise ValueError(
        f"No se encontro sheet de resumen. Sheets disponibles: {wb.sheetnames}"
    )


# ── Formato nuevo (2026-06-11) ───────────────────────────────────────────────
# Wansoft reemplazó SalesByBranch por ConsolidatedSalesMasterReport.
# El XLSX nuevo concentra todo en el sheet "Resumen de Ventas" con secciones
# horizontales: Ventas por hora / platillo / grupo / tipo de pago / usuario...

def _find_new_resumen_ws(wb: openpyxl.Workbook):
    for name in wb.sheetnames:
        if name.strip().lower() == "resumen de ventas":
            return wb[name]
    raise ValueError(
        f"No se encontro sheet 'Resumen de Ventas'. Sheets: {wb.sheetnames}"
    )


def _find_section_header(ws, header_text: str, max_row: int = 15):
    """Localiza la celda cuyo valor exacto es header_text. Returns (row, col) o None."""
    target = header_text.strip().lower()
    for row in ws.iter_rows(min_row=1, max_row=max_row):
        for c in row:
            if isinstance(c.value, str) and c.value.strip().lower() == target:
                return c.row, c.column
    return None


def _parse_usuario_section(wb: openpyxl.Workbook):
    """Sección 'Ventas por usuario' del formato nuevo → mismas keys que legacy."""
    ws = _find_new_resumen_ws(wb)
    pos = _find_section_header(ws, "Usuario")
    if pos is None:
        raise ValueError("No se encontro seccion 'Ventas por usuario' (header 'Usuario')")
    hrow, hcol = pos

    headers = {}
    for row in ws.iter_rows(min_row=hrow, max_row=hrow, min_col=hcol, max_col=hcol + 10):
        for i, c in enumerate(row):
            if c.value is not None:
                headers[str(c.value).strip().lower()] = i

    def col(name):
        if name not in headers:
            raise ValueError(
                f"Columna '{name}' no encontrada en seccion usuario. Headers: {list(headers)}"
            )
        return headers[name]

    i_total = col("total")          # con IVA — cuadra con la app de Wansoft
    i_personas = col("no. personas")
    i_promedio = col("promedio por persona")

    rows = []
    for row in ws.iter_rows(
        min_row=hrow + 1, min_col=hcol, max_col=hcol + 10, values_only=True
    ):
        usuario = row[0]
        if usuario is None or str(usuario).strip() == "":
            break  # fin de la sección
        u = str(usuario).strip()
        if u.lower() in ("total", "totales", "gran total"):
            continue
        total_val = row[i_total]
        personas_val = row[i_personas]
        if total_val is None or personas_val is None:
            continue
        rows.append({
            "mesero": u,
            "total": float(total_val),
            "personas": int(float(personas_val)),
            "promedio": float(row[i_promedio] or 0),
        })
    return rows


def _parse_sheet(wb: openpyxl.Workbook):
    """Parse mesero rows — intenta formato legacy, luego el nuevo."""
    try:
        return _parse_sheet_legacy(wb)
    except ValueError:
        return _parse_usuario_section(wb)


def _parse_sheet_legacy(wb: openpyxl.Workbook):
    """Parse resumen sheet (formato viejo SalesByBranch). Returns list of dicts."""
    ws = _find_resumen_sheet(wb)

    # Find header row — look for cell containing "Mesero" as header
    header_row = None
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=15, values_only=False), start=1
    ):
        for cell in row:
            if cell.value and str(cell.value).strip().lower() == "mesero":
                header_row = row_idx
                break
        if header_row:
            break

    if header_row is None:
        raise ValueError("No se encontro la fila de headers con 'Mesero'")

    # Read headers and build column map
    headers_raw = [c.value for c in ws[header_row]]
    col_map = {}
    for i, h in enumerate(headers_raw):
        if h is None:
            continue
        hl = str(h).strip().lower()
        if hl == "mesero":
            col_map["mesero"] = i
        elif hl == "total":
            col_map["total"] = i
        elif "persona" in hl and "promedio" not in hl:
            col_map["personas"] = i
        elif "promedio" in hl:
            col_map["promedio"] = i

    required = {"mesero", "total", "personas", "promedio"}
    missing = required - set(col_map.keys())
    if missing:
        raise ValueError(
            f"Columnas faltantes: {missing}. Headers: {headers_raw}"
        )

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        mesero = row[col_map["mesero"]]
        if mesero is None or str(mesero).strip() == "":
            continue
        mesero_str = str(mesero).strip()
        if mesero_str.lower() in ("total", "totales", "gran total"):
            continue
        total_val = row[col_map["total"]]
        personas_val = row[col_map["personas"]]
        promedio_val = row[col_map["promedio"]]
        if total_val is None or personas_val is None:
            continue
        rows.append({
            "mesero": mesero_str,
            "total": float(total_val),
            "personas": int(float(personas_val)),
            "promedio": float(promedio_val),
        })

    return rows


def _extract_date(wb: openpyxl.Workbook):
    """Extract report date from metadata rows."""
    # Scan first 6 rows of every sheet for date patterns
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
            for cell in row:
                if cell is None:
                    continue
                if isinstance(cell, (datetime, date)):
                    return cell if isinstance(cell, date) else cell.date()
                s = str(cell)
                # "Reporte del: 2026-05-09 al 2026-05-09"
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"):
                    for match in re.findall(
                        r"\d{1,4}[/-]\d{1,2}[/-]\d{2,4}", s
                    ):
                        try:
                            return datetime.strptime(match, fmt).date()
                        except ValueError:
                            continue
    return None


def format_message(xlsx_path: str, report_type: str = "cierre") -> str:
    """Parse XLSX and return WhatsApp-formatted message.

    report_type: "cierre" (end-of-day complete) or "avance" (mid-day partial).
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    rows = _parse_sheet(wb)
    report_date = _extract_date(wb)
    wb.close()

    if report_date is None:
        m = re.search(r"(\d{4})-?(\d{2})-?(\d{2})", Path(xlsx_path).stem)
        if m:
            report_date = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        else:
            report_date = date.today()

    # Filter out non-mesero rows
    filtered = [r for r in rows if not _is_excluded(r["mesero"])]
    if not filtered:
        return "Sin datos de meseros para este dia.", {}

    # Sort by promedio descending
    filtered.sort(key=lambda r: r["promedio"], reverse=True)

    # Totals across filtered meseros
    total_dia = sum(r["total"] for r in filtered)
    personas_dia = sum(r['personas'] for r in filtered)
    general_avg = total_dia / personas_dia if personas_dia else 0.0

    # Date formatting
    dd = f"{report_date.day:02d}"
    mes = MESES_ES[report_date.month]
    date_str = f"{dd} {mes}"

    # Header by report type
    if report_type == "avance":
        header = f"☀️ *AMALAY · Avance del día · {date_str} (3pm)*"
    else:
        header = f"🌙 *AMALAY · Cierre {date_str}*"

    # Build message
    lines = [header, ""]
    lines.append(f"Ticket promedio del día: {_fmt_money(general_avg)}")
    lines.append("")

    # Top 3
    top3 = filtered[:3]
    for i, r in enumerate(top3):
        medal = MEDALS[i]
        lines.append(
            f"{medal} {_short_name(r['mesero'])} {_fmt_money(r['promedio'])} "
            f"({r['personas']} personas)"
        )

    # Rest
    rest = filtered[3:]
    if rest:
        lines.append("")
        lines.append("Resto:")
        for j in range(0, len(rest), 2):
            pair = rest[j:j + 2]
            parts = [
                f"{_short_name(r['mesero'])} {_fmt_money(r['promedio'])} ({r['personas']})"
                for r in pair
            ]
            lines.append("- " + " · ".join(parts))

    # Top volumen note
    personas_list = [r['personas'] for r in filtered]
    median_personas = statistics.median(personas_list)
    threshold = median_personas * 1.5
    high_vol = [r for r in filtered if r['personas'] > threshold]
    if high_vol:
        top_vol = max(high_vol, key=lambda r: r['personas'])
        lines.append("")
        lines.append(
            f"🏆 Top mesas: {_short_name(top_vol['mesero'])} ({top_vol['personas']})"
        )

    lines.append("")
    lines.append(f"Total día: {_fmt_money(total_dia)} · {personas_dia} personas")
    lines.append("")
    lines.append("_by Fullsite ✨_")

    # Footnote by report type
    if report_type == "avance":
        lines.append("_Datos parciales — al cierre llegará el resumen completo_")
    else:
        lines.append("_Promedios mezclan turnos brunch (mañana) y cafecito (tarde) — v2 separará por turno_")

    msg = "\n".join(lines)
    agg = {
        "total_dia": total_dia,
        "personas_dia": personas_dia,
        "ticket_promedio": general_avg,
        "meseros_top": [
            {"nombre": _short_name(r["mesero"]), "total": r["total"],
             "personas": r["personas"], "promedio": r["promedio"]}
            for r in filtered
        ],
    }
    return msg, agg


DETALLE_SHEET = "Ventas por mesero por grupo"

POSTRES_GRUPOS = {"DESSERTS", "ICE CREAM"}


def _find_detalle_sheet(wb: openpyxl.Workbook):
    """Find the 'Ventas por mesero por grupo' sheet."""
    if DETALLE_SHEET in wb.sheetnames:
        return wb[DETALLE_SHEET]
    for name in wb.sheetnames:
        if "grupo" in name.lower() and "mesero" in name.lower() and "resumen" not in name.lower():
            return wb[name]
    # Fallback: first sheet (original Wansoft layout)
    return wb[wb.sheetnames[0]]


def _parse_platillos_section(wb: openpyxl.Workbook):
    """Sección 'Ventas por platillo / artículo' del formato nuevo.

    El export nuevo NO trae desglose platillo×mesero — mesero queda "".
    """
    ws = _find_new_resumen_ws(wb)
    pos = _find_section_header(ws, "Nombre Platillo/Artículo")
    if pos is None:
        raise ValueError("No se encontro seccion 'Ventas por platillo / artículo'")
    hrow, hcol = pos

    headers = {}
    for row in ws.iter_rows(min_row=hrow, max_row=hrow, min_col=hcol, max_col=hcol + 5):
        for i, c in enumerate(row):
            if c.value is not None:
                headers[str(c.value).strip().lower()] = i
    if "grupo" not in headers or "cantidad" not in headers:
        raise ValueError(f"Headers inesperados en seccion platillos: {list(headers)}")

    rows = []
    for row in ws.iter_rows(
        min_row=hrow + 1, min_col=hcol, max_col=hcol + 5, values_only=True
    ):
        nombre = row[0]
        if nombre is None or str(nombre).strip() == "":
            break
        cantidad = row[headers["cantidad"]]
        if cantidad is None:
            continue
        rows.append({
            "mesero": "",
            "grupo": str(row[headers["grupo"]] or "").strip().upper(),
            "platillo": str(nombre).strip().upper(),
            "cantidad": int(float(cantidad)),
        })
    return rows


def _parse_detalle_sheet(wb: openpyxl.Workbook):
    """Parse platillo rows — intenta formato legacy, luego el nuevo."""
    try:
        return _parse_detalle_legacy(wb)
    except ValueError:
        return _parse_platillos_section(wb)


def _parse_detalle_legacy(wb: openpyxl.Workbook):
    """Parse granular platillo sheet (formato viejo). Returns list of row dicts."""
    ws = _find_detalle_sheet(wb)

    # Find header row with Mesero + Platillo + Cantidad
    header_row = None
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=15, values_only=False), start=1
    ):
        vals = [str(c.value).strip().lower() if c.value else "" for c in row]
        if "mesero" in vals and "platillo" in vals and "cantidad" in vals:
            header_row = row_idx
            break

    if header_row is None:
        raise ValueError("No se encontro header row con Mesero/Platillo/Cantidad")

    headers_raw = [c.value for c in ws[header_row]]
    col_map = {}
    for i, h in enumerate(headers_raw):
        if h is None:
            continue
        hl = str(h).strip().lower()
        if hl == "mesero":
            col_map["mesero"] = i
        elif hl == "grupo":
            col_map["grupo"] = i
        elif hl == "platillo":
            col_map["platillo"] = i
        elif hl == "cantidad":
            col_map["cantidad"] = i

    required = {"mesero", "grupo", "platillo", "cantidad"}
    missing = required - set(col_map.keys())
    if missing:
        raise ValueError(f"Columnas faltantes en detalle: {missing}")

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        mesero = row[col_map["mesero"]]
        if mesero is None or str(mesero).strip() == "":
            continue
        platillo = str(row[col_map["platillo"]] or "").strip()
        grupo = str(row[col_map["grupo"]] or "").strip()
        cantidad = row[col_map["cantidad"]]
        if cantidad is None:
            continue
        rows.append({
            "mesero": str(mesero).strip(),
            "grupo": grupo.upper(),
            "platillo": platillo.upper(),
            "cantidad": int(float(cantidad)),
        })

    return rows


def _wrap_mesero_line(items: list[tuple[str, int]], max_width: int = 80) -> list[str]:
    """Wrap 'Name N · Name N · ...' lines at ~max_width chars."""
    lines = []
    current = []
    current_len = 0
    for name, count in items:
        entry = f"{name} {count}"
        entry_len = len(entry) + (3 if current else 0)  # " · " separator
        if current and current_len + entry_len > max_width:
            lines.append(" · ".join(current))
            current = [entry]
            current_len = len(entry)
        else:
            current.append(entry)
            current_len += entry_len
    if current:
        lines.append(" · ".join(current))
    return lines


def format_platillos_message(xlsx_path: str, report_type: str = "cierre") -> str:
    """Parse XLSX Sheet 1 and return platillo detail WhatsApp message."""
    from collections import defaultdict

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    rows = _parse_detalle_sheet(wb)
    report_date = _extract_date(wb)
    wb.close()

    if report_date is None:
        m = re.search(r"(\d{4})-?(\d{2})-?(\d{2})", Path(xlsx_path).stem)
        if m:
            report_date = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        else:
            report_date = date.today()

    # Date formatting
    dd = f"{report_date.day:02d}"
    mes = MESES_ES[report_date.month]
    date_str = f"{dd} {mes}"

    # Aggregations
    hh_total = 0
    chilaquiles_total = 0
    bakery_by_mesero = defaultdict(int)
    postres_by_mesero = defaultdict(int)

    for r in rows:
        platillo = r["platillo"]
        grupo = r["grupo"]
        mesero = r["mesero"]
        cant = r["cantidad"]

        if "HALF" in platillo and "COMBO" in platillo:
            hh_total += cant
        if "CHILAQUILES" in platillo:
            chilaquiles_total += cant
        if grupo == "BAKERY" and not _is_excluded(mesero):
            bakery_by_mesero[mesero] += cant
        if grupo in POSTRES_GRUPOS and not _is_excluded(mesero):
            postres_by_mesero[mesero] += cant

    # Sort and filter zeros
    bakery_sorted = sorted(
        [(m, c) for m, c in bakery_by_mesero.items() if c > 0],
        key=lambda x: x[1], reverse=True,
    )
    postres_sorted = sorted(
        [(m, c) for m, c in postres_by_mesero.items() if c > 0],
        key=lambda x: x[1], reverse=True,
    )

    # Header
    if report_type == "avance":
        header = f"📊 *AMALAY · Detalle platillos · {date_str} (3pm)*"
    else:
        header = f"📊 *AMALAY · Detalle platillos {date_str}*"

    lines = [header, ""]
    lines.append(f"🥗 *H&H Combo:* {hh_total} vendidos")
    lines.append(f"🌮 *Chilaquiles:* {chilaquiles_total} vendidos")

    bakery_total = sum(c for _, c in bakery_sorted)
    postres_total = sum(c for _, c in postres_sorted)
    has_mesero = any(r["mesero"] for r in rows)

    if has_mesero:
        # Pan dulce por mesero
        lines.append("")
        lines.append("🥐 *Pan dulce por mesero:*")
        bakery_items = [(_short_name(m), c) for m, c in bakery_sorted]
        for line in _wrap_mesero_line(bakery_items):
            lines.append(line)
        lines.append(f"Total: {bakery_total} piezas")

        # Postres por mesero
        lines.append("")
        lines.append("🍰 *Postres por mesero:*")
        postres_items = [(_short_name(m), c) for m, c in postres_sorted]
        for line in _wrap_mesero_line(postres_items):
            lines.append(line)
        lines.append(f"Total: {postres_total} piezas")
    else:
        # Formato nuevo: Wansoft ya no desglosa platillo×mesero
        lines.append("")
        lines.append(f"🥐 *Pan dulce:* {bakery_total} piezas")
        lines.append(f"🍰 *Postres:* {postres_total} piezas")

    lines.append("")
    lines.append("_by Fullsite ✨_")

    msg = "\n".join(lines)

    # Aggregate platillo-level data by grupo for platillos_top
    from collections import Counter
    grupo_totals = Counter()
    for r in rows:
        if not _is_excluded(r["mesero"]):
            grupo_totals[r["grupo"]] += r["cantidad"]
    platillos_top = [
        {"nombre": g, "total": c}
        for g, c in grupo_totals.most_common()
    ]

    agg = {
        "chilaquiles_total": chilaquiles_total,
        "half_half_total": hh_total,
        "platillos_top": platillos_top,
    }
    return msg, agg


if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(description="Parse Wansoft XLSX → WhatsApp message")
    ap.add_argument("xlsx", help="Path to XLSX file")
    ap.add_argument("--type", dest="report_type", choices=["cierre", "avance"],
                    default="cierre", help="Report type (default: cierre)")
    args = ap.parse_args()

    if not Path(args.xlsx).exists():
        print(f"Error: archivo no encontrado: {args.xlsx}", file=sys.stderr)
        sys.exit(1)
    msg, _ = format_message(args.xlsx, report_type=args.report_type)
    print(msg)
    print()
    plat_msg, _ = format_platillos_message(args.xlsx, report_type=args.report_type)
    print(plat_msg)
